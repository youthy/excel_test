# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import xlrd
import re
import glob
from openpyxl import Workbook
files = glob.glob('./*.xlsx')
try: 
    files.remove('.\\output.xlsx')
    files.remove('./output.xlsx')
except:
    pass
#origin = 'origin.xlsx'
output = 'output.xlsx'
#fd = xlrd.open_workbook(origin)
wb = Workbook()
ws = wb.active
key_dict = {}
extract_words = [0, '手机号码', '身份证', '居住地址', '户籍地址', '居住年限', '现工作年限','车型', '车 型', '实际销售价格','贷款年限', '贷款金额',
                 '月供', '车架号', '发票单位', '开户行', '帐号', '开户名', '借款合同号', '二手车']
for fn in files:
    fd = xlrd.open_workbook(fn)
    for sheet in fd.sheets():
        if sheet.nrows < 2:
            continue
        row_to_write = ws.max_row + 1
        for index, item in enumerate(sheet.col_values(0)):
            if item == '':
                continue
            if isinstance(item, float):
                
                key = index,
                
                val0 = xlrd.xldate_as_tuple(item, fd.datemode)
                
                y = val0[0]
                m = val0[1]
                d = val0[2]
                val = "{0}/{1}/{2}".format(y, m, d)
               
            elif isinstance(item, int) or isinstance(item, float):
                key = index
                val = item
            else:
                k, *v = re.split(': |： |:|：', item)
                if v == []:
                    if index < 10:
                        key = index
                        val = k
                    elif k =="二手车":
                        key = k
                        val = "是"
                    else:
                        key = k
                        val = ''
                else:
                    key = k
                    val = v[0]
                if key == '手机号码':
                    val, *_ = val.split('(')
            if key in extract_words:
                col_n = key_dict.get(key, None)
                if col_n == None:
                    col_n = len(key_dict) + 1
                    key_dict[key] = col_n
                ws.cell(row = row_to_write, column = col_n).value = val
                if key == '居住地址':
                    m = re.search(r'^.*省|^[^省]*市', val)
                    colnn = key_dict.get('省份', None)
                    if colnn == None:
                        key_dict['省份'] = col_n + 1
                        colnn = col_n + 1
                    if m:
                        province = m.group()
                    else:
                        province = ''
                    ws.cell(row = row_to_write, column = colnn).value = province
    for k, v in key_dict.items():
        ws.cell(row = 1, column = v).value=str(k)
wb.save(filename = output)
            
    
    
