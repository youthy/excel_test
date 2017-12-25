# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""
import xlrd
import re
import glob
from openpyxl import Workbook
files = glob.glob('./*.xlsx')
try: 
    files.remove('.\\output.xlsx')
    files.remove('./output.xlsx')
except:
    pass
#origin = 'origin.xlsx'
output = 'output.xlsx'
#fd = xlrd.open_workbook(origin)
wb = Workbook()
ws = wb.active
key_dict = {}
for fn in files:
    fd = xlrd.open_workbook(fn)
    for sheet in fd.sheets():
        if sheet.nrows < 2:
            continue
        row_to_write = ws.max_row + 1
        for index, item in enumerate(sheet.col_values(0)):
            if item == '':
                continue
            if isinstance(item, float):
                
                key = index,
                
                val0 = xlrd.xldate_as_tuple(item, fd.datemode)
                
                y = val0[0]
                m = val0[1]
                d = val0[2]
                val = "{0}/{1}/{2}".format(y, m, d)
               
            elif isinstance(item, int) or isinstance(item, float):
                key = index
                val = item
            else:
                k, *v = re.split(': |： |:|：', item)
                if v == []:
                    if index < 10:
                        key = index
                        val = k
                    else:
                        key = k
                        val = ''
                else:
                    key = k
                    val = v[0]
            col_n = key_dict.get(key, None)
            if col_n == None:
                col_n = len(key_dict) + 1
                key_dict[key] = col_n
            ws.cell(row = row_to_write, column = col_n).value = val
    for k, v in key_dict.items():
        ws.cell(row = 1, column = v).value=str(k)
wb.save(filename = output)
            
    
    
